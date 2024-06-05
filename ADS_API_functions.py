#Author: Cole Howell
#Purpose: Functions to utilize the ADS API
#Completed: 1/19/2023
#Updates: 1/23/23: Fixed an issue with the updating of the I/I map where the rain total dictionary was a nested
# dictionary making the subtraction of values not work

import openpyxl
import requests
import json
import datetime as dt
import os
import pandas as pd
from array import array
from arcgis.gis import GIS
from openpyxl import load_workbook
import numpy as np
from scipy.stats import mannwhitneyu
import matplotlib.pyplot as plt
import high_tide_api_functions as htt
import time
import configparser


##################################|Functions for Running Totals|###########################################
#configure script with file containing sensitive information
def config():
    config = configparser.ConfigParser()
    config.read('C:\EsriTraining\PythonGP\Scripts\wadc_dev_config.ini')

    #file contains username, password, and api token/parameter information
    gis = config['GIS']
    ads = config['ads']
    htt = config['high_tide']

    return gis, ads, htt


#get yesterday
def get_yesterday():
    return dt.datetime.now().date() - dt.timedelta(days=1)


#location of the log book
def log_book_location():
    user = os.environ['USERNAME']
    directory = "C:/Users/" + user + "/WADC Dropbox/Water Authority/SEWER BASINS/"

    out_xlsx = directory + "Sewer Meter Total Daily Volume Log 2023.xlsx"

    return out_xlsx


#raises a value error when the server sends back a bad response, accepts the code as an argument
def raise_server_error(code):
    if code >= 400:
        print(code)
        raise ValueError('Error accessing the server for the ADS API.')


#gets the names and ids of all entities stored in ads
def request_entities(ads):

    response = requests.get("https://api.adsprism.com/api/Entity", ads['x-ads-dev'])
    print(response.content)


#gets the active meters from the server and loads them into a python object and returns the python object
def get_active_meters(ads):
    params = {
        "x-ads-dev": ads['x-ads-dev'],
        "showInactives": False
    }
    response = requests.get("https://api.adsprism.com/api/Locations", params)

    raise_server_error(response.status_code)
    active = json.loads(response.content)

    return active


#returns a list of active meter ids
def get_meter_ids(active_meters):
    list_of_ids = []
    for meter in active_meters:
        list_of_ids.append(meter.get('id'))
    return list_of_ids


#returns a dictionary where the ids are the keys and the location names are the values
def get_meter_names(active_meters):
    name_dict = {}
    for meter in active_meters:
        name_dict[meter.get('id')] = meter.get('name')

    return name_dict


#gets telemetry for the total flow from the active meters in a 24 hour period from midnight yesterday to just before
#midnight today, accepts a list of meter ids as an argument
def get_active_telemetry(ids, ads):
    params = {
        "x-ads-dev": ads['x-ads-dev'],
        "locationId": array("i", ids),
        "entityId": array("i", [3338, 3339, 2123]),
        "start": str(get_yesterday()) + "T00:00:00",
        "end": str(get_yesterday()) + "T23:59:59",
        "locationGroupId": None
    }
    response = requests.get("https://api.adsprism.com/api/Telemetry", params)
    #raises an error if the server returns a bad status
    raise_server_error(response.status_code)

    telemetry = json.loads(response.content)
    return telemetry


#takes the first and last values in each triton telemetry and calculates the daily total (kgal) for it,
# then assigns the value to a dictionary by id number and returns the dictionary
def calculate_daily_totals(telemetries, names):
    triton_dictionary = {}
    temp_list = [[]]
    temp_list2 = []
    i = 0
    j = 0
    for telemetry in telemetries:
        #print(telemetry['locationId'])
        if telemetry['locationId'] == 98 or telemetry['locationId'] == 102 or telemetry['locationId'] == 104:
            continue
        elif telemetry['locationId'] == 92 or telemetry['locationId'] == 108:
            continue
            #initial = telemetry['entityData'][2]['data'][0]['reading']
            #final = telemetry['entityData'][2]['data'][len(telemetry['entityData'][0]['data']) - 1]['reading']
        elif telemetry['locationId'] == 78 or telemetry['locationId'] == 83:
            continue
        else:
            #these if and elif statements handle the data from tritons with two sensors connected
            #schrader heights
            if telemetry['locationId'] == 97:
                temp_list2.append([])
                initial2 = telemetry['entityData'][1]['data'][0]['reading']
                final2 = telemetry['entityData'][1]['data'][len(telemetry['entityData'][1]['data']) - 1]['reading']
                daily_total_2 = final2 - initial2
                temp_list2[j].append(daily_total_2)
                triton_dictionary[names[98]] = temp_list2[j]
                j += 1
            #tennsco rd
            elif telemetry['locationId'] == 99:
                temp_list2.append([])
                if telemetry['entityData'][1]['data']:
                    initial2 = telemetry['entityData'][1]['data'][0]['reading']
                    #print(len(telemetry['entityData'][0]['data']))
                    final2 = telemetry['entityData'][1]['data'][len(telemetry['entityData'][1]['data']) - 1]['reading']
                    daily_total_2 = final2 - initial2
                    temp_list2[j].append(daily_total_2)
                else:
                    daily_total_2 = 0.0
                    temp_list2[j].append(daily_total_2)
                triton_dictionary[names[102]] = temp_list2[j]
                j += 1
            #cowan rd
            elif telemetry['locationId'] == 103:
                temp_list2.append([])
                if telemetry['entityData'][1]['data']:
                    initial2 = telemetry['entityData'][1]['data'][0]['reading']
                    final2 = telemetry['entityData'][1]['data'][len(telemetry['entityData'][1]['data']) - 1]['reading']
                    daily_total_2 = final2 - initial2
                    temp_list2[j].append(daily_total_2)
                else:
                    daily_total_2 = 0.0
                    temp_list2[j].append(daily_total_2)
                triton_dictionary[names[104]] = temp_list2[j]
                j += 1
            #lewis hollow, west is now picked up by the
            elif telemetry['locationId'] == 105:
                temp_list2.append([])
                if telemetry['entityData'][1]['data']:
                    initial2 = telemetry['entityData'][1]['data'][0]['reading']
                    final2 = telemetry['entityData'][1]['data'][len(telemetry['entityData'][1]['data']) - 1]['reading']
                    daily_total_2 = final2 - initial2
                    temp_list2[j].append(daily_total_2)
                else:
                    daily_total_2 = 0.0
                    temp_list2[j].append(daily_total_2)
                triton_dictionary['Lewis_Hollow_West'] = temp_list2[j]
                j += 1
            #print(telemetry['locationId'])
            #print(telemetry['entityData'][0]['data'][0])
            if telemetry['entityData'][0]['data']:
                initial = telemetry['entityData'][0]['data'][0]['reading']
                final = telemetry['entityData'][0]['data'][len(telemetry['entityData'][0]['data']) - 1]['reading']
            else:
                initial = 0.0
                final = 0.0

        daily_total = final - initial
        temp_list.append([])
        temp_list[i].append(daily_total)
        triton_dictionary[names[telemetry['locationId']]] = temp_list[i]
        #print(triton_dictionary)
        i += 1
    return triton_dictionary


#log yesterday's volume totals in an excel file, accept the triton data dictionary and the site names dictionary
def log_totals(tritons, out_xlsx):

    new_dict = {
        'Date': get_yesterday()
    }
    new_dict.update(tritons)
    #print(new_dict)
    #this needs the dictionaries to have lists as values for some reason
    temp_df = pd.DataFrame.from_dict(new_dict)

    #if the excel file already exists append the new data otherwise make a new file
    if os.path.exists(out_xlsx):
        old_df = pd.read_excel(out_xlsx, engine='openpyxl')
        #print(overall_df)
        #append new data to old data and send to excel
        writer = pd.ExcelWriter(out_xlsx, engine='openpyxl', mode='a', if_sheet_exists='replace')

        #checks the main dataframe if all the columns from the new data exist and if not adds an empty list to it
        for column in temp_df:
            #print(column)
            if column in old_df.columns:
                continue
            else:
                empty_list = []
                for i in range(0, len(old_df.index)):
                    empty_list.append(0.0)
                    i += 1
                old_df[column] = empty_list

        overall_df = pd.concat([old_df, temp_df], ignore_index=True)
        #print(overall_df)
        overall_df.to_excel(writer, sheet_name='Daily Total Gallons (kgal)', index=False)
        #alter size of columns
        #for column in overall_df:
        #    column_length = max(overall_df[column].astype(str).map(len).max(), len(column))
        #   col_idx = overall_df.columns.get_loc(column)
        #  writer.sheets['Daily Total Gallons (kgal)'].set_column(col_idx, col_idx, column_length)
    else:
        #send the new data to the file
        writer = pd.ExcelWriter(out_xlsx, engine='xlsxwriter')
        temp_df.to_excel(writer, sheet_name='Daily Total Gallons (kgal)', index=False)
        #alter size of columns
        for column in temp_df:
            column_length = max(temp_df[column].astype(str).map(len).max(), len(column))
            col_idx = temp_df.columns.get_loc(column)
            writer.sheets['Daily Total Gallons (kgal)'].set_column(col_idx, col_idx, column_length)

    writer.save()
    #writer.close()


#edits the log book stored on the enterprise service
def log_in_gis(telemetries, tritons, g):
    rain_total = 0
    # add up all the rain measurements in inches
    for telemetry in telemetries:
        if telemetry['locationId'] == 92:
            for i in range(0, len(telemetry['entityData'][2]['data'])):
                rain_total += telemetry['entityData'][2]['data'][i]['reading']

    gis = GIS("https://esriapps1.esriwadc.com/portal", g['username'], g['password'])
    portal_item = gis.content.get('6b829b25aeed491ea8309911c5491914')
    table_layer = portal_item.tables[0]
    yesterday = dt.datetime.now() - dt.timedelta(days=1)
    yesterday = time.mktime(yesterday.timetuple())

    for triton in tritons:
        add = {"attributes":{
            "date": yesterday*10**(3),
            "ads_meter": triton,
            "flow": float(tritons[triton][0]),
            "dickson_rain": rain_total
        }}
        table_layer.edit_features(adds=[add])
        # print(add)


#log the rain total for yesterday
def log_rain(telemetries, out_xlsx):

    rain_total = 0
    #add up all the rain measurements in inches
    for telemetry in telemetries:
        if telemetry['locationId'] == 92:
            for i in range(0, len(telemetry['entityData'][2]['data'])):
                rain_total += telemetry['entityData'][2]['data'][i]['reading']

    temp_list = [[]]
    temp_list[0].append(rain_total)
    rain_dict = {
        "Date": get_yesterday(),
        "Dickson_Rain": temp_list[0]
    }

    temp_df = pd.DataFrame(data=rain_dict)
    wb = load_workbook(out_xlsx, read_only=True)

    # if the excel file already exists append the new data otherwise make a new file
    if 'Daily Rain Total (in)' in wb.sheetnames:
        old_df = pd.read_excel(out_xlsx, sheet_name='Daily Rain Total (in)', engine='openpyxl')
        # print(overall_df)
        # append new data to old data and send to excel
        writer = pd.ExcelWriter(out_xlsx, engine='openpyxl', mode='a', if_sheet_exists='replace')

        # checks the main dataframe if all the columns from the new data exist and if not adds an empty list to it
        for column in temp_df:
            # print(column)
            if column in old_df.columns:
                continue
            else:
                empty_list = []
                for i in range(0, len(old_df.index)):
                    empty_list.append(0)
                    i += 1
                old_df[column] = empty_list

        overall_df = pd.concat([old_df, temp_df], ignore_index=True)
        # print(overall_df)
        overall_df.to_excel(writer, sheet_name='Daily Rain Total (in)', index=False)
        # alter size of columns
        #for column in overall_df:
        #    column_length = max(overall_df[column].astype(str).map(len).max(), len(column))
        #    col_idx = overall_df.columns.get_loc(column)
        #    writer.sheets['Daily Rain Total (in)'].set_column(col_idx, col_idx, column_length)
    else:
        # send the new data to the file
        writer = pd.ExcelWriter(out_xlsx, engine='openpyxl', mode='a')
        temp_df.to_excel(writer, sheet_name='Daily Rain Total (in)', index=False)
        # alter size of columns
        #for column in temp_df:
        #    column_length = max(temp_df[column].astype(str).map(len).max(), len(column))
        #    col_idx = temp_df.columns.get_loc(column)
        #    writer.sheets['Daily Rain Total (in)'].set_column(col_idx, col_idx, column_length)

    writer.save()
    #writer.close()


#format the cell size in the log book
def format_logbook(out_xlsx):

    wb = openpyxl.load_workbook(out_xlsx)

    for sheet in wb.worksheets:
        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value
    wb.save(out_xlsx)


#send the totals to arcgis
def update_gis(tritons, g):
    gis = GIS("https://esriapps1.esriwadc.com/portal", g['username'], g['password'])
    portal_item = gis.content.get('6e0558dc0c654674aa97c323d3bc8c0a')
    meter_layer = portal_item.layers[0]
    meter_fset = meter_layer.query()
    meter_features = meter_fset.features

    #loop through the meter locations in gis and update the daily total attributes
    for meter in meter_features:
        edit = meter
        if edit.attributes['id_1'] in tritons:
            edit.attributes['daily_total'] = tritons[edit.attributes['id_1']][0]
            #print(edit.attributes['daily_total'])
        else:
            edit.attributes['daily_total'] = None
        meter_layer.edit_features(updates=[edit])


##################################|Functions for the I/I map|#########################################################
#calculates the running average baseline flow through each meter by averaging volume measurements on days without rain
def running_average(out_xlsx):
    flow_df = pd.read_excel(out_xlsx, sheet_name='Daily Total Gallons (kgal)', engine='openpyxl')
    flow_df.set_index('Date')
    rain_df = pd.read_excel(out_xlsx, sheet_name='Daily Rain Total (in)', engine='openpyxl')

    flow_df['Rain'] = rain_df['Dickson_Rain']

    #selects the rows where there was no significant rain recorded
    flow_df = flow_df.loc[flow_df['Rain'] < 0.1]

    #deletes the rain column since it's no longer needed
    flow_df = flow_df.drop(columns='Rain')

    #flow_df.clip(lower=0)
    flow_df.replace(0, np.nan, inplace=True)

    #calculates the mean value of the baseline flows through the meters
    flow_series = flow_df.mean(numeric_only=True, skipna=True)

    average = flow_series.to_dict()
    # print(average)

    return average


#get the average rain flow per zone
def avg_rain(out_xlsx):
    flow_df = pd.read_excel(out_xlsx, sheet_name='Daily Total Gallons (kgal)', engine='openpyxl')
    flow_df.set_index('Date')
    rain_df = pd.read_excel(out_xlsx, sheet_name='Daily Rain Total (in)', engine='openpyxl')

    flow_df['Rain'] = rain_df['Dickson_Rain']

    # selects the rows where there was no significant rain recorded
    flow_df = flow_df.loc[flow_df['Rain'] > 0.09]

    # deletes the rain column since it's no longer needed
    flow_df = flow_df.drop(columns='Rain')

    # flow_df.clip(lower=0)
    flow_df.replace(0, np.nan, inplace=True)

    # calculates the mean value of the baseline flows through the meters
    flow_series = flow_df.mean(numeric_only=True, skipna=True)

    average = flow_series.to_dict()
    # print(average)

    return average


#get the flow readings from the last rain
def last_rain(out_xlsx):
    flow_df = pd.read_excel(out_xlsx, sheet_name='Daily Total Gallons (kgal)', engine='openpyxl')
    flow_df.set_index('Date')
    rain_df = pd.read_excel(out_xlsx, sheet_name='Daily Rain Total (in)', engine='openpyxl')

    flow_df['Rain'] = rain_df['Dickson_Rain']
    # print(flow_df)
    # selects the rows where there was rain recorded
    flow_df = flow_df.loc[flow_df['Rain'] > 0.09]
    # print(flow_df)
    #gets the last one
    flow_df = flow_df.iloc[-1:]

    #deletes the rain column
    flow_df = flow_df.drop(columns='Rain')
    #flow_df.replace(0, np.nan, inplace=True)

    # print(flow_df)

    rained = flow_df.to_dict('list')
    # print(rained)

    return rained


#dictionary for flow corrections based on mass balance equations for the zones
def mass_balance():
    balance = {
        'Gentry_Circle': ['Gentry_Circle'],
        'Schrader_Heights_': ['Schrader_Heights_', 'Gentry_Circle'],
        'Schrader_Heights_(2)': ['Schrader_Heights_(2)', 'Tennsco_Rd', 'Tennsco_Rd(2)'],
        'Printwood': ['Printwood', 'Dal-Tile_Quartz'],
        'Sherry_Ln': ['Sherry_Ln'],
        'First_Federal': ['First_Federal', 'Sherry_Ln', 'Schrader_Heights_', 'Schrader_Heights_(2)', 'Printwood'],
        'Sherwin_Williams': ['Sherwin_Williams'],
        'Lewis_Hollow_West': ['Lewis_Hollow_West'],
        'Lewis_Hollow_East': ['Lewis_Hollow_East'],
        'Firestone': ['Firestone', 'Sherwin_Williams', 'Lewis_Hollow_East', 'Lewis_Hollow_East(2)'],
        'NAPA': ['NAPA'],
        'Cowan_Rd': ['Cowan_Rd', 'NAPA', 'Firestone', 'First_Federal'],
        'Cowan_Rd(2)': ['Cowan_Rd(2)'],
        'Tennsco_Rd': ['Tennsco_Rd'],
        'Dal-Tile_Quartz': ['Dal-Tile_Quartz'],
        'Dal-Tile': ['Dal-Tile']
    }
    return balance


def gis_mass_balance(date, g, h):

    #include the mass balance schema
    balance = mass_balance()
    balance['Tennsco_Rd(2)'] = ['Tennsco_Rd(2)']

    #access the gis
    gis = GIS("https://esriapps1.esriwadc.com/portal", g['username'], g['password'])

    #get the meter flow and rain table
    flow_portal_item = gis.content.get('6b829b25aeed491ea8309911c5491914')
    flow_table_layer = flow_portal_item.tables[0]
    flow_set = flow_table_layer.query()
    table_rows = flow_set.features

    #get the mass balance table
    mass_item = gis.content.get('3f31ce8162614dc7ae773362f611b9ad')
    mass_table = mass_item.tables[0]

    table_rows = [r for r in table_rows if dt.datetime.fromtimestamp(r.attributes['date']*10**(-3)).date() == date]
    # print(table_rows)
    #loop through the flow log table
    for row in table_rows:
        #make a list of the appropriate meters to use for the math based on date and the balance dictionary
        rows_date = [r for r in table_rows if r.attributes['date'] == row.attributes['date']
                     and r.attributes['ads_meter'] in balance[row.attributes['ads_meter']]]
        # print(rows_date)
        m_bal = 0
        #loop through the rows_date list
        for r in rows_date:
            if r.attributes['ads_meter'] == row.attributes['ads_meter']:
                m_bal += row.attributes['flow']

                #the first federal zone needs to use the high tide data for the sumiden ls for the mass balance
                if row.attributes['ads_meter'] == 'First_Federal' and row.attributes['date'] <= 1681275600000 \
                        and row.attributes['dickson_rain'] < 0.1:
                    m_bal += 130
                elif row.attributes['ads_meter'] == 'First_Federal' and row.attributes['date'] <= 1681275600000 \
                        and row.attributes['dickson_rain'] >= 0.1:
                    m_bal += 330
                if row.attributes['ads_meter'] == 'First_Federal' and row.attributes['date'] > 1681275600000:
                    # type(row.attributes['date'])
                    sumiden = htt.get_sumiden_stats(dt.datetime.fromtimestamp(row.attributes['date']*10**(-3)).date(), h)
                    m_bal += ((sumiden[2]['data'][len(sumiden[2]['data'])-1]['reading'] - sumiden[2]['data'][0]['reading']))*10**(-3)
            else:
                if row.attributes['ads_meter'] == 'First_Federal':
                    # print(dt.datetime.fromtimestamp(row.attributes['date']*10**(-3)).date())
                    sumiden = htt.get_sumiden_stats(dt.datetime.fromtimestamp(row.attributes['date']*10**(-3)).date(), h)
                    influent_1 = (sumiden[0]['data'][len(sumiden[0]['data']) - 1]['reading'] - sumiden[0]['data'][0][
                        'reading'])
                    influent_2 = (sumiden[1]['data'][len(sumiden[1]['data']) - 1]['reading'] - sumiden[1]['data'][0][
                        'reading'])
                    m_bal -= (influent_1 + influent_2)*10**(-3)
                    sherry = [r for r in rows_date if r.attributes['ads_meter'] == 'Sherry_Ln']
                    m_bal -= sherry[0].attributes['flow']
                    break
                else:
                    m_bal -= r.attributes['flow']
        # m_row.attributes['mass_balance_flow'] = m_bal
        # mass_table.edit_features(updates=[m_row])

        add = {'attributes':{
            'date': row.attributes['date'],
            'sub_basin': row.attributes['ads_meter'],
            'mass_balance_flow': m_bal,
            'rain': row.attributes['dickson_rain'],
            'meters': balance[row.attributes['ads_meter']]
        }}
        # print(add)

        mass_table.edit_features(adds=[add])


#update the infiltration zones hosted feature layer
def infiltration_zones(average, rained, disc_rain, p_values, g):
    gis = GIS("https://esriapps1.esriwadc.com/portal", g['username'], g['password'])
    portal_item = gis.content.get('b6f0cf98d7bf436591261de49bac4a43')
    infil_layer = portal_item.layers[0]
    infil_fset = infil_layer.query()
    infil_features = infil_fset.features

    balance = mass_balance()

    #loop through the infiltration features
    for infil in infil_features:
        edit = infil
        if edit.attributes['meter'] in average:
            if edit.attributes['meter'] == "Tennsco_Rd(2)":
                continue
            #store the raw average baseline in gis
            edit.attributes['avg_baseline'] = average[edit.attributes['meter']]

            #store the average rain in appropriate zone
            edit.attributes['avg_rain'] = rained[edit.attributes['meter']]
            #store the last rain event totals in gis
            edit.attributes['last_rain'] = disc_rain[edit.attributes['meter']][0]
            edit.attributes['p_value'] = p_values[edit.attributes['meter']]

            #mass balance average
            zone_corrected_avg = 0
            for i in range(0, len(balance[edit.attributes['meter']])):
                if i == 0:
                    zone_corrected_avg = average[balance[edit.attributes['meter']][i]]
                    i += 1
                else:
                    zone_corrected_avg -= average[balance[edit.attributes['meter']][i]]
                    i += 1
            if edit.attributes['meter'] == 'First_Federal':
                zone_corrected_avg += (213.0336 + 207.7056)
            edit.attributes['mass_balance_avg'] = zone_corrected_avg

            #mass balance rain
            zone_corrected_rain = 0
            for i in range(0, len(balance[edit.attributes['meter']])):
                if i == 0:
                    zone_corrected_rain = rained[balance[edit.attributes['meter']][i]]
                    i += 1
                else:
                    zone_corrected_rain -= rained[balance[edit.attributes['meter']][i]]
                    i += 1

            #Much of the water from the first federal leg is sent through suminen lift station, I'm adding the average
            #lift station flow to these values to account for it
            if edit.attributes['meter'] == 'First_Federal':
                zone_corrected_rain += (244.8432 + 246.2112)
            edit.attributes['mass_balance_rain'] = zone_corrected_rain

            #calculate the i/i and find the ratio of it to the total flow during the last rain for each meter
            ini = edit.attributes['mass_balance_rain'] - edit.attributes['mass_balance_avg']
            if edit.attributes['last_rain'] == 0:
                percent_ini = 0
                edit.attributes['percent_ini'] = None
            else:
                percent_ini = ini / edit.attributes['mass_balance_rain']
                edit.attributes['percent_ini'] = percent_ini * 100.0

            #check the ratio and assign the infiltration risk values to proper colors
            if percent_ini <= 0.15:
                if edit.attributes['mass_balance_rain'] < 0 or edit.attributes['mass_balance_avg'] < 0 or \
                        (percent_ini < 0 and abs(percent_ini*100) > 2.0):
                    edit.attributes['infiltration_risk'] = 3
                else:
                    edit.attributes['infiltration_risk'] = 0
            elif percent_ini > 0.15 and (percent_ini <= 0.45):
                if edit.attributes['mass_balance_rain'] < 0 or edit.attributes['mass_balance_avg'] < 0:
                    edit.attributes['infiltration_risk'] = 3
                else:
                    edit.attributes['infiltration_risk'] = 1
            else:
                if edit.attributes['mass_balance_rain'] < 0 or edit.attributes['mass_balance_avg'] < 0:
                    edit.attributes['infiltration_risk'] = 3
                else:
                    edit.attributes['infiltration_risk'] = 2

        else:
            edit.attributes['avg_baseline'] = None
            edit.attributes['last_rain'] = None
        infil_layer.edit_features(updates=[edit])


#performs a mann-whitney u-test on the data
def mann_whitney_u(out_xlsx):
    flow_df = pd.read_excel(out_xlsx, sheet_name='Daily Total Gallons (kgal)', engine='openpyxl')
    flow_df.set_index('Date')
    rain_df = pd.read_excel(out_xlsx, sheet_name='Daily Rain Total (in)', engine='openpyxl')

    flow_df['Rain'] = rain_df['Dickson_Rain']
    # print(flow_df)
    # selects the rows where there was rain recorded
    rain_df = flow_df.loc[flow_df['Rain'] > 0.09]
    base_df = flow_df.loc[flow_df['Rain'] < 0.1]

    rain_df = rain_df.drop(columns='Rain')
    base_df = base_df.drop(columns='Rain')

    rain_df = rain_df.drop(columns='Date')
    base_df = base_df.drop(columns='Date')

    rain_df.replace(0, np.nan, inplace=True)
    base_df.replace(0, np.nan, inplace=True)

    p_value_dict = {}
    for column in base_df:
        result = mannwhitneyu(base_df[column], rain_df[column])
        p_value_dict[column] = result[1]

    return p_value_dict


#makes histograms from all of the data in the log book separated by rain and baseline
def histogram_generator(out_xlsx):
    flow_df = pd.read_excel(out_xlsx, sheet_name='Daily Total Gallons (kgal)', engine='openpyxl')
    flow_df.set_index('Date')
    rain_df = pd.read_excel(out_xlsx, sheet_name='Daily Rain Total (in)', engine='openpyxl')

    flow_df['Rain'] = rain_df['Dickson_Rain']
    # print(flow_df)
    # selects the rows where there was rain recorded
    rain_df = flow_df.loc[flow_df['Rain'] > 0.09]
    base_df = flow_df.loc[flow_df['Rain'] < 0.1]

    rain_df = rain_df.drop(columns='Rain')
    base_df = base_df.drop(columns='Rain')

    rain_df = rain_df.drop(columns='Date')
    base_df = base_df.drop(columns='Date')

    rain_df.replace(0, np.nan, inplace=True)
    base_df.replace(0, np.nan, inplace=True)

    plt.rcParams["figure.figsize"] = [7.00, 3.5]
    plt.rcParams["figure.autolayout"] = True

    for column in base_df:
        plt.figure()
        plt.hist(base_df[column])
        plt.savefig('C:/Users/chowell/WADC Dropbox/Cole Howell/PC/Documents/Sewer Infiltration/Histograms/Baseline/'
                    + str(column) + ".png")

        plt.figure()
        plt.hist(rain_df[column])
        plt.savefig('C:/Users/chowell/WADC Dropbox/Cole Howell/PC/Documents/Sewer Infiltration/Histograms/Rain/'
                    + str(column) + ".png")


#calculates the mass balance for the last rain event
def last_rain_balance(book, g, h):
    balance = mass_balance()
    rain = last_rain(book)
    mb = {}
    mb_rain = 0

    #get the runtimes for the sumiden lift station and average them together
    sumiden = htt.get_sumiden_stats(rain['Date'][0], h)
    influent_1 = (sumiden[0]['data'][len(sumiden[0]['data']) - 1]['reading'] - sumiden[0]['data'][0]['reading'])
    influent_2 = (sumiden[1]['data'][len(sumiden[1]['data']) - 1]['reading'] - sumiden[1]['data'][0]['reading'])
    discharge = (sumiden[2]['data'][len(sumiden[2]['data'])-1]['reading'] - sumiden[2]['data'][0]['reading'])
    # avg_runtime = np.average(runtimes)

    #loop through the balance dictionary and perform the mass balance calculation
    for key in balance:
        for i, lst in enumerate(balance[key]):
            if i == 0:
                mb_rain = rain[lst][0]
            else:
                if key == 'First_Federal' and i > 1:
                    mb_rain += (discharge - influent_1 - influent_2)*10**(-3)
                    break
                mb_rain -= rain[lst][0]
        # if key == 'First_Federal':
        #     mb_rain += 0.4*60*avg_runtime
        mb[key] = mb_rain

    gis = GIS("https://esriapps1.esriwadc.com/portal", g['username'], g['password'])
    portal_item = gis.content.get('b6f0cf98d7bf436591261de49bac4a43')
    infil_layer = portal_item.layers[0]
    infil_fset = infil_layer.query()
    infil_features = infil_fset.features

    for infil in infil_features:
        edit = infil
        if edit.attributes['meter'] in mb:
            edit.attributes['last_rain_mb'] = mb[edit.attributes['meter']]
            edit.attributes['last_rain_infil'] = (100 * (edit.attributes['last_rain_mb'] -
                                                         edit.attributes['mass_balance_avg'])/
                                                  edit.attributes['last_rain_mb'])
        else:
            edit.attributes['last_rain_mb'] = None
            edit.attributes['last_rain_infil'] = None
        infil_layer.edit_features(updates=[edit])
